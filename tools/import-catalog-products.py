from __future__ import annotations

import hashlib
import io
import json
import re
import shutil
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from PIL import Image, ImageOps
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
PUBLIC = ROOT / "public"
ASSET_DIR = PUBLIC / "assets" / "catalog"
DATA_FILE = PUBLIC / "data" / "site-data.json"
REPORT_FILE = ROOT / "data" / "catalog-import-report.json"

SOURCES = {
    "benz_oem": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/奔驰图册.xlsx"),
    "bmw_oem": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/宝马图册.xlsx"),
    "bmw_headlights": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/配齐 宝马oe 改装大灯.xlsx"),
    "benz_headlights": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/配齐 奔驰oe 改装大灯.xlsx"),
    "tail_lights": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/配齐 奔驰宝马原车尾灯 改装尾灯.xlsx"),
    "tianju": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/天聚产品目录.xlsx"),
    "lima": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/msg/file/2026-06/LIMA-Lightting Catalog.xlsx"),
    "exhaust_pdf": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/产品目录.pdf"),
    "akd_bmw": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/AKD BMW Catalogue 2026.pdf"),
    "akd_benz": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/AKD Benz Catalogue 2026.pdf"),
    "akd_audi": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/AKD AUDI Catalogue 2026.pdf"),
    "akd_vw": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/AKD VW Catalogue.pdf"),
    "akd_porsche": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/AKD Porsche Catalogue.pdf"),
    "linlur": Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_ewzmcm5vnxo322_c7cd/temp/drag/南通玲路汽车有限公司.pdf"),
}

BRAND_CN = {
    "宝马": "BMW",
    "奔驰": "Mercedes-Benz",
    "奥迪": "Audi",
    "保时捷": "Porsche",
    "大众": "Volkswagen",
    "特斯拉": "Tesla",
}

MODEL_CN = {
    "高尔夫": "Golf",
    "卡宴": "Cayenne",
    "帕拉梅拉": "Panamera",
    "帕纳梅拉": "Panamera",
    "马卡": "Macan",
    "途锐": "Touareg",
    "途观": "Tiguan",
    "帕萨特": "Passat",
    "朗逸": "Lavida",
    "迈腾": "Magotan",
    "速腾": "Sagitar",
    "捷达": "Jetta",
    "宝来": "Bora",
    "思域": "Civic",
    "雅阁": "Accord",
    "飞度": "Fit",
    "凯美瑞": "Camry",
    "汉兰达": "Highlander",
    "锐志": "Reiz",
    "埃尔法": "Alphard",
    "普拉多": "Prado",
    "陆巡": "Land Cruiser",
    "蒙迪欧": "Mondeo",
    "野马": "Mustang",
    "福克斯": "Focus",
    "揽胜": "Range Rover",
    "极光": "Evoque",
    "发现": "Discovery",
    "明锐": "Octavia",
    "速派": "Superb",
}

PLACEHOLDERS = {
    "Automotive Lighting": "assets/live/category-lighting.png",
    "Tail Lights": "assets/live/category-lighting.png",
    "Exhaust Systems": "assets/live/category-exhaust.png",
    "Body Kits": "assets/live/category-body-kits.png",
    "OEM Replacement Parts": "assets/live/logo.jpg",
    "Catalog Reference": "assets/live/category-lighting.png",
}


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_part(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\n", " / ").strip())


def remove_cjk(value: str) -> str:
    value = re.sub(r"[\u3400-\u9fff]+", " ", value)
    return text(value)


def has_cjk(value: str) -> bool:
    return bool(re.search(r"[\u3400-\u9fff]", value))


def slugify(value: str, fallback: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value[:90] or fallback


def stable_id(*parts: str) -> str:
    raw = "|".join(parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:10]


def year_en(value: str) -> str:
    value = text(value)
    value = value.replace("年款", "").replace("款", "").replace("年", "")
    value = value.replace("至", "-").replace("—", "-").replace("~", "-")
    value = re.sub(r"(?<!\d)(\d{2})-(\d{2})(?!\d)", lambda m: f"20{m.group(1)}-20{m.group(2)}", value)
    return value


def detect_brand(value: str, fallback: str = "") -> str:
    combined = value.lower()
    for cn, en in BRAND_CN.items():
        if cn in value:
            return en
    if "mercedes" in combined or "benz" in combined:
        return "Mercedes-Benz"
    if "bmw" in combined:
        return "BMW"
    if "audi" in combined:
        return "Audi"
    if "porsche" in combined or "panamera" in combined or "cayenne" in combined or "macan" in combined:
        return "Porsche"
    if "volkswagen" in combined or "golf" in combined or "vw" in combined:
        return "Volkswagen"
    if "tesla" in combined or "tsl" in combined:
        return "Tesla"
    return fallback


def product_type_from_text(value: str, default: str = "Automotive Part") -> tuple[str, str]:
    lower = value.lower()
    if any(x in value for x in ["尾灯", "后尾灯"]) or "tail" in lower or "rear lamp" in lower or "rearlamp" in lower:
        return "Tail Lights", "Tail Light Assembly"
    if any(x in value for x in ["大灯", "前照灯", "前灯"]) or "headlight" in lower or "headlamp" in lower:
        return "Automotive Lighting", "Headlight Assembly"
    if "exhaust" in lower or "downpipe" in lower or "catback" in lower or "排气" in value:
        return "Exhaust Systems", "Performance Exhaust System"
    if "body kit" in lower or "bumper" in lower or "spoiler" in lower or "包围" in value:
        return "Body Kits", "Body Kit"
    return default, default


def features_from_text(value: str) -> list[str]:
    checks = [
        ("LED", ["LED", "led"]),
        ("halogen", ["卤素", "halogen"]),
        ("xenon", ["氙气", "xenon"]),
        ("laser", ["激光", "laser"]),
        ("matrix", ["矩阵", "matrix"]),
        ("sequential turn signal", ["流水", "跑马", "dynamic turn", "sequential"]),
        ("DRL", ["日行", "DRL", "drl"]),
        ("projector lens", ["透镜", "projector"]),
        ("facelift conversion", ["老改新", "改新"]),
        ("plug-and-play upgrade", ["直接对插", "plug"]),
        ("smoked housing", ["熏黑", "smoked"]),
        ("RS style", ["RS", "rs"]),
        ("AMG style", ["AMG", "amg"]),
    ]
    result = []
    for label, keys in checks:
        if any(k in value for k in keys):
            result.append(label)
    return result


def model_from_text(value: str, brand: str = "") -> str:
    v = value.replace("\n", " ")
    for cn, en in MODEL_CN.items():
        if cn in v:
            suffixes = " ".join(re.findall(r"\b(?:W|S|C|X|G|F|E|A|B|V|R|MK|958|957|970|971|9Y0|G8X|G2[023]|F\d{2}|E\d{2}|\d{2,4})[A-Z0-9.]*\b", v, flags=re.I))
            return text(f"{en} {suffixes}")
    patterns = [
        r"\b(?:W|S|C|X|G|F|E|A|B|V|R|GLA|GLC|GLE|GLS|CLA|CLS|ML|GL|G|M|RS|S|Q|TT|F|G|E)\d{2,4}[A-Z]?\b",
        r"\b(?:A|B|C|E|S|G|V|ML|GL|GLA|GLC|GLE|GLS|CLA|CLS)\s?class\b",
        r"\b(?:A|Q|S|RS)\d(?:L)?\b",
        r"\b(?:Golf|Polo|Passat|Tiguan|Touareg|Cayenne|Panamera|Macan|Boxster|718|911)\b",
        r"\b(?:1|2|3|4|5|6|7|8)\s?(?:Series|SER)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, v, flags=re.I)
        if match:
            return match.group(0).replace("  ", " ")
    if brand:
        v = v.replace(brand, "")
    for cn, en in BRAND_CN.items():
        v = v.replace(cn, "")
    for cn, en in MODEL_CN.items():
        v = v.replace(cn, en)
    v = re.sub(r"headlights?|taillights?|rear ?lamp|front|for|大灯|尾灯|总成|改装|升级", " ", v, flags=re.I)
    return remove_cjk(v)[:40]


def english_title(brand: str, model: str, year: str, product_type: str, raw: str, part_no: str = "") -> str:
    feature = ", ".join(features_from_text(raw)[:3])
    pieces = [brand, model, feature, product_type, year]
    title = " ".join([p for p in pieces if p])
    title = remove_cjk(title)
    if not title:
        title = f"{brand or 'Automotive'} {product_type}"
    if part_no and not has_cjk(part_no) and len(title) < 90:
        title = f"{title} ({part_no})"
    return title[:180]


def save_image(data: bytes, rel_stem: str) -> str | None:
    try:
        image = Image.open(io.BytesIO(data))
        image = ImageOps.exif_transpose(image)
        if image.mode not in ("RGB", "RGBA"):
            image = image.convert("RGBA" if "A" in image.getbands() else "RGB")
        image.thumbnail((900, 900), Image.Resampling.LANCZOS)
        if image.mode == "RGBA":
            bg = Image.new("RGB", image.size, "white")
            bg.paste(image, mask=image.getchannel("A"))
            image = bg
        out = ASSET_DIR / f"{rel_stem}.webp"
        out.parent.mkdir(parents=True, exist_ok=True)
        image.save(out, "WEBP", quality=82, method=6)
        return str(out.relative_to(PUBLIC))
    except Exception:
        return None


def images_by_row(path: Path, source_key: str) -> dict[str, dict[int, str]]:
    wb = load_workbook(path, read_only=False, data_only=True)
    by_sheet: dict[str, dict[int, str]] = defaultdict(dict)
    for ws in wb.worksheets:
        for index, image in enumerate(getattr(ws, "_images", []), start=1):
            try:
                row = image.anchor._from.row + 1
                data = image._data()
                rel = save_image(data, f"{source_key}/{slugify(ws.title, 'sheet')}-{row}-{index}")
                if rel and row not in by_sheet[ws.title]:
                    by_sheet[ws.title][row] = rel
            except Exception:
                continue
    return by_sheet


def wps_cell_images(path: Path, source_key: str) -> dict[str, str]:
    result: dict[str, str] = {}
    with zipfile.ZipFile(path) as zf:
        try:
            cell_xml = zf.read("xl/cellimages.xml")
            rel_xml = zf.read("xl/_rels/cellimages.xml.rels")
        except KeyError:
            return result
        rel_root = ET.fromstring(rel_xml)
        rels = {node.attrib["Id"]: node.attrib["Target"] for node in rel_root}
        root = ET.fromstring(cell_xml)
        ns = {
            "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }
        for pic in root.findall(".//xdr:pic", ns):
            name_node = pic.find(".//xdr:cNvPr", ns)
            blip = pic.find(".//a:blip", ns)
            if name_node is None or blip is None:
                continue
            image_id = name_node.attrib.get("name", "")
            embed = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "")
            target = rels.get(embed, "")
            if not image_id or not target:
                continue
            try:
                data = zf.read("xl/" + target)
                rel = save_image(data, f"{source_key}/{image_id.lower()}")
                if rel:
                    result[image_id] = rel
            except Exception:
                continue
    return result


def make_product(source: str, title: str, category: str, brand: str = "", model: str = "", year: str = "", raw: str = "", image: str = "", **extra: Any) -> dict[str, Any]:
    title = remove_cjk(re.sub(r"\([^)]*[\u3400-\u9fff][^)]*\)", " ", title))
    model = remove_cjk(model)
    year = remove_cjk(year)
    uid = f"{source}-{stable_id(title, brand, model, year, raw)}"
    slug = slugify(f"{brand}-{model}-{title}-{uid[-6:]}", uid)
    category = category or "OEM Replacement Parts"
    local_image = image or PLACEHOLDERS.get(category, "assets/live/logo.jpg")
    description = extra.pop("description", "")
    if not description:
        description = f"Request a quote for {title}. Confirm year, model, trim, LHD/RHD, connector, MOQ, lead time, packaging, and shipping before ordering."
    return {
        "id": uid,
        "slug": slug,
        "title": title,
        "url": "",
        "price": extra.pop("price", ""),
        "compareAt": "",
        "status": extra.pop("status", "RFQ"),
        "source": source,
        "localImage": local_image,
        "category": category,
        "brand": brand,
        "model": model,
        "yearRange": year,
        "productType": extra.pop("productType", product_type_from_text(title)[1]),
        "partNumbers": extra.pop("partNumbers", []),
        "side": extra.pop("side", ""),
        "material": extra.pop("material", ""),
        "moq": extra.pop("moq", ""),
        "rawSourceName": raw[:220],
        "description": description,
        "features": extra.pop("features", features_from_text(raw + " " + title)),
        **extra,
    }


def import_existing() -> list[dict[str, Any]]:
    data = json.loads(DATA_FILE.read_text())
    result = []
    for i, product in enumerate(data.get("products", [])):
        if not str(product.get("source", "")).startswith("https://cowinmotors.com"):
            continue
        p = dict(product)
        p.setdefault("id", f"current-{i}")
        p.setdefault("slug", slugify(p.get("title", ""), f"current-{i}"))
        p.setdefault("brand", detect_brand(p.get("title", "")))
        p.setdefault("model", model_from_text(p.get("title", ""), p.get("brand", "")))
        p.setdefault("yearRange", "")
        p.setdefault("productType", product_type_from_text(p.get("title", ""), p.get("category", ""))[1])
        p.setdefault("partNumbers", [])
        p.setdefault("description", f"Request a quote for {p.get('title', 'this automotive product')} with fitment confirmation and export shipping support.")
        p.setdefault("features", features_from_text(p.get("title", "")))
        result.append(p)
    return result


def import_oem_general(path: Path, source: str, brand: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        part_no, name, model, origin, *_ = list(row) + [""] * 5
        part_no = clean_part(part_no)
        name = clean_part(name)
        model = clean_part(model)
        if not part_no or not name:
            continue
        title = english_title(brand, model, "", "OEM Replacement Part", name, part_no)
        products.append(make_product(source, title, "OEM Replacement Parts", brand, model, "", name, partNumbers=[part_no], origin=clean_part(origin), productType="OEM Replacement Part"))
    return products


def import_bmw_headlights(path: Path) -> list[dict[str, Any]]:
    image_map = images_by_row(path, "bmw-headlights")
    wb = load_workbook(path, read_only=True, data_only=True)
    products = []
    for ws in wb.worksheets:
        if ws.title not in {"大灯", "改装", "emak"}:
            continue
        current_series = ""
        current_chassis = ""
        for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            values = list(row) + [""] * 10
            if ws.title == "改装":
                _, series, raw_name, _, _, fitment, remark, flag = values[:8]
                current_series = clean_part(series) or current_series
                raw = clean_part(raw_name)
                if not raw:
                    continue
                year = year_en(clean_part(fitment))
                model = model_from_text(f"{current_series} {raw} {fitment}", "BMW")
                title = english_title("BMW", model, year, "Headlight Upgrade Kit", raw + " " + clean_part(remark))
                products.append(make_product("配齐 宝马oe 改装大灯.xlsx", title, "Automotive Lighting", "BMW", model, year, raw, image_map.get(ws.title, {}).get(idx, ""), fitment=clean_part(fitment), notes=clean_part(remark), productType="Headlight Upgrade Kit"))
            else:
                _, series, chassis, raw_name, _, year, parts, attr, remark, note2 = values[:10]
                current_series = clean_part(series) or current_series
                current_chassis = clean_part(chassis) or current_chassis
                raw = clean_part(raw_name)
                if not raw:
                    continue
                yr = year_en(clean_part(year))
                model = current_chassis or model_from_text(f"{current_series} {raw}", "BMW")
                title = english_title("BMW", model, yr, "Headlight Assembly", raw + " " + clean_part(attr), clean_part(parts).split(" / ")[0])
                products.append(make_product("配齐 宝马oe 改装大灯.xlsx", title, "Automotive Lighting", "BMW", model, yr, raw, image_map.get(ws.title, {}).get(idx, ""), partNumbers=[p.strip() for p in clean_part(parts).split("/") if p.strip()], productType="Headlight Assembly", notes="; ".join(filter(None, [clean_part(attr), clean_part(remark), clean_part(note2)]))))
    return products


def import_benz_headlights(path: Path) -> list[dict[str, Any]]:
    image_map = images_by_row(path, "benz-headlights")
    wb = load_workbook(path, read_only=True, data_only=True)
    products = []
    for ws in wb.worksheets:
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = list(row) + [""] * 8
            part_no, _, raw_name, model, inner, outer, weight = values[:7]
            part_no = clean_part(part_no)
            raw = clean_part(raw_name)
            model = clean_part(model)
            if not raw and not part_no:
                continue
            yr = year_en(model)
            m = model_from_text(model + " " + part_no, "Mercedes-Benz")
            product_type = "Headlight Upgrade Kit" if ws.title == "改装" else "Headlight Assembly"
            title = english_title("Mercedes-Benz", m or model, yr, product_type, raw + " " + model, part_no)
            products.append(make_product("配齐 奔驰oe 改装大灯.xlsx", title, "Automotive Lighting", "Mercedes-Benz", m or model, yr, raw, image_map.get(ws.title, {}).get(idx, ""), partNumbers=[part_no] if part_no else [], productType=product_type, carton=clean_part(inner or outer), weight=clean_part(weight)))
    return products


def import_tail_lights(path: Path) -> list[dict[str, Any]]:
    image_map = images_by_row(path, "tail-lights")
    wb = load_workbook(path, read_only=True, data_only=True)
    products = []
    for ws in wb.worksheets:
        brand = "Mercedes-Benz" if ws.title == "Benz" else "BMW"
        current_class = ""
        current_name = ""
        current_year = ""
        current_image = ""
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            cls, name, year, _, oem, side, price, qty, *_ = list(row) + [""] * 9
            current_class = clean_part(cls) or current_class
            current_name = clean_part(name) or current_name
            current_year = year_en(clean_part(year) or current_year)
            current_image = image_map.get(ws.title, {}).get(idx, current_image)
            oem = clean_part(oem)
            if not current_name or not oem:
                continue
            model = model_from_text(current_class + " " + current_name, brand)
            title = english_title(brand, model, current_year, "Tail Light Assembly", current_name, oem)
            products.append(make_product("配齐 奔驰宝马原车尾灯 改装尾灯.xlsx", title, "Tail Lights", brand, model, current_year, current_name, current_image, partNumbers=[oem], side=clean_part(side), price=f"RMB {clean_part(price)}" if clean_part(price) else "", moq=clean_part(qty), productType="Tail Light Assembly"))
    return products


def import_tianju(path: Path) -> list[dict[str, Any]]:
    img_by_id = wps_cell_images(path, "tianju")
    wb = load_workbook(path, read_only=True, data_only=True)
    products = []
    formula_re = re.compile(r'DISPIMG\("([^"]+)"')
    for ws in wb.worksheets:
        category = "Tail Lights" if "尾" in ws.title else "Automotive Lighting"
        product_type = "Tail Light Assembly" if category == "Tail Lights" else "Headlight Assembly"
        for row in ws.iter_rows(min_row=2, values_only=False):
            raw = clean_part(row[0].value)
            brand = detect_brand(clean_part(row[1].value), detect_brand(raw))
            yr = year_en(clean_part(row[2].value))
            if not raw or not brand:
                continue
            formula = str(row[3].value or "")
            match = formula_re.search(formula)
            image = img_by_id.get(match.group(1), "") if match else ""
            model = model_from_text(raw, brand)
            title = english_title(brand, model, yr, product_type, raw)
            products.append(make_product("天聚产品目录.xlsx", title, category, brand, model, yr, raw, image, productType=product_type))
    return products


def import_lima(path: Path) -> list[dict[str, Any]]:
    image_map = images_by_row(path, "lima-lighting")
    wb = load_workbook(path, read_only=True, data_only=True)
    products = []
    for ws in wb.worksheets:
        brand = {"Benz": "Mercedes-Benz", "VW": "Volkswagen"}.get(ws.title, ws.title)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            values = list(row) + [""] * 12
            code, _, model_raw, year, color, size, weight, *rest = values[:12]
            model_raw = clean_part(model_raw)
            if not model_raw:
                continue
            category, ptype = product_type_from_text(model_raw, "Automotive Lighting")
            yr = year_en(clean_part(year))
            model = model_from_text(model_raw, brand)
            raw = " ".join(filter(None, [model_raw, clean_part(color), clean_part(rest[0] if rest else "")]))
            title = english_title(brand, model or model_raw, yr, ptype, raw, clean_part(code))
            products.append(make_product("LIMA-Lightting Catalog.xlsx", title, category, brand, model or model_raw, yr, raw, image_map.get(ws.title, {}).get(idx, ""), partNumbers=[clean_part(code)] if clean_part(code) else [], color=clean_part(color), size=clean_part(size), weight=clean_part(weight), productType=ptype))
    return products


def import_exhaust_pdf(path: Path) -> list[dict[str, Any]]:
    reader = PdfReader(path)
    text_blob = "\n".join((page.extract_text() or "") for page in reader.pages)
    text_blob = re.sub(r"\s+", " ", text_blob)
    pattern = re.compile(r"(DC-[A-Z]+-\d+)\s+(.+?)(?=\s+DC-[A-Z]+-\d+|$)")
    products = []
    for part_no, body in pattern.findall(text_blob):
        body = body.strip()
        if len(body) < 8:
            continue
        material = "SS304 stainless steel" if "304" in body or "stainless" in body.lower() else ""
        moq_match = re.search(r"(\d+)\s*Sets?", body, flags=re.I)
        moq = f"{moq_match.group(1)} sets" if moq_match else ""
        model = re.sub(r"\b304\b|stainles+s? steel|stainless steel|\d+ Sets?", " ", body, flags=re.I)
        model = text(model)[:80]
        brand = detect_brand(model, "BMW" if "BMW" in model.upper() else "")
        title = english_title(brand, model, "", "Performance Exhaust System", body, part_no)
        products.append(make_product("产品目录.pdf", title, "Exhaust Systems", brand, model, "", body, "assets/live/category-exhaust.png", partNumbers=[part_no], material=material, moq=moq, productType="Performance Exhaust System"))
    return products


def import_pdf_visual_refs(path: Path, source: str, brand: str, category_hint: str) -> list[dict[str, Any]]:
    reader = PdfReader(path)
    products = []
    for index, page in enumerate(reader.pages, start=1):
        rel = ""
        try:
            images = sorted(page.images, key=lambda img: len(img.data), reverse=True)
            if images:
                rel = save_image(images[0].data, f"pdf-visuals/{slugify(source, 'pdf')}-page-{index}")
        except Exception:
            rel = ""
        title = f"{brand} {category_hint} Catalog Visual Reference Page {index}"
        products.append(make_product(source, title, "Catalog Reference", brand, "", "", title, rel, productType="Catalog Visual Reference", description=f"Visual catalog reference page from {source}. Send the page number, vehicle model, and target product type for quote confirmation."))
    return products


def dedupe(products: list[dict[str, Any]]) -> list[dict[str, Any]]:
    seen = set()
    out = []
    slug_counts: dict[str, int] = defaultdict(int)
    for product in products:
        key = (product.get("brand", ""), product.get("model", ""), product.get("yearRange", ""), product.get("productType", ""), tuple(product.get("partNumbers", [])), product.get("title", ""))
        if key in seen:
            continue
        seen.add(key)
        base = product["slug"]
        slug_counts[base] += 1
        if slug_counts[base] > 1:
            product["slug"] = f"{base}-{slug_counts[base]}"
        out.append(product)
    return out


def main() -> None:
    if ASSET_DIR.exists():
        shutil.rmtree(ASSET_DIR)
    products = import_existing()
    stats = {"existing": len(products)}
    importers = [
        ("benz_oem", lambda: import_oem_general(SOURCES["benz_oem"], "奔驰图册.xlsx", "Mercedes-Benz")),
        ("bmw_oem", lambda: import_oem_general(SOURCES["bmw_oem"], "宝马图册.xlsx", "BMW")),
        ("bmw_headlights", lambda: import_bmw_headlights(SOURCES["bmw_headlights"])),
        ("benz_headlights", lambda: import_benz_headlights(SOURCES["benz_headlights"])),
        ("tail_lights", lambda: import_tail_lights(SOURCES["tail_lights"])),
        ("tianju", lambda: import_tianju(SOURCES["tianju"])),
        ("lima", lambda: import_lima(SOURCES["lima"])),
        ("exhaust_pdf", lambda: import_exhaust_pdf(SOURCES["exhaust_pdf"])),
        ("akd_bmw", lambda: import_pdf_visual_refs(SOURCES["akd_bmw"], "AKD BMW Catalogue 2026.pdf", "BMW", "Lighting")),
        ("akd_benz", lambda: import_pdf_visual_refs(SOURCES["akd_benz"], "AKD Benz Catalogue 2026.pdf", "Mercedes-Benz", "Lighting")),
        ("akd_audi", lambda: import_pdf_visual_refs(SOURCES["akd_audi"], "AKD AUDI Catalogue 2026.pdf", "Audi", "Lighting")),
        ("akd_vw", lambda: import_pdf_visual_refs(SOURCES["akd_vw"], "AKD VW Catalogue.pdf", "Volkswagen", "Lighting")),
        ("akd_porsche", lambda: import_pdf_visual_refs(SOURCES["akd_porsche"], "AKD Porsche Catalogue.pdf", "Porsche", "Lighting")),
        ("linlur", lambda: import_pdf_visual_refs(SOURCES["linlur"], "南通玲路汽车有限公司.pdf", "Porsche", "Headlight, Tail Light and Body Kit")),
    ]
    for key, fn in importers:
        before = len(products)
        try:
            products.extend(fn())
            stats[key] = len(products) - before
        except Exception as error:
            stats[key] = f"ERROR: {error}"
    products = dedupe(products)
    products.sort(key=lambda p: (p.get("category", ""), p.get("brand", ""), p.get("model", ""), p.get("title", "")))
    data = json.loads(DATA_FILE.read_text())
    data["products"] = products
    DATA_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    (PUBLIC / "data" / "live-products.json").write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    REPORT_FILE.parent.mkdir(parents=True, exist_ok=True)
    report = {
        "totalProducts": len(products),
        "stats": stats,
        "categoryCounts": dict(sorted(defaultdict(int, ((cat, sum(1 for p in products if p.get("category") == cat)) for cat in {p.get("category", "") for p in products})).items())),
        "brandCounts": dict(sorted(defaultdict(int, ((brand, sum(1 for p in products if p.get("brand") == brand)) for brand in {p.get("brand", "") for p in products})).items())),
    }
    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
